from openpyxl import Workbook, load_workbook
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
target_ar = load_workbook(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_arg v1.0.xlsx'))
target_cmms = load_workbook(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_cmm v1.0.xlsx'))
target_ecm = load_workbook(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_ecm v1.0.xlsx'))


def get_mappings():

    source_wb = load_workbook(os.path.join(BASE_DIR, r'tables\FR.XLSX'))
    print(source_wb.get_sheet_names())

    sheet_names = source_wb.get_sheet_names()[1:]

    asset_register = source_wb[sheet_names[0]]
    cmms = source_wb[sheet_names[1]]
    ecm = source_wb[sheet_names[2]]

    target_ar_sheet = target_ar[target_ar.get_sheet_names()[-1]]
    target_cmms_sheet = target_cmms[target_cmms.get_sheet_names()[-1]]
    target_ecm_sheet = target_ecm[target_ecm.get_sheet_names()[-1]]

    mappings = dict()
    mappings[asset_register] = target_ar_sheet
    mappings[cmms] = target_cmms_sheet
    mappings[ecm] = target_ecm_sheet

    print(mappings)
    return mappings


def save_wbs():
    target_ar.save(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_arg v1.0.xlsx'))
    target_cmms.save(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_cmm v1.0.xlsx'))
    target_ecm.save(os.path.join(BASE_DIR, r'tables\gal - asset data platform  fr list_ecm v1.0.xlsx'))


mappings = get_mappings()
for k, v in mappings.items():
    source = k
    target = v

    source_description = dict()
    true_description = dict()
    source_resp = dict()

    for col in source.iter_cols(min_row=4, min_col=2, max_col=2):
        for cell in col:
            value = cell.value.strip()
            if len(value) == 5:
                source_description[value] = ""
                true_description[value] = ""
                source_resp[value] = ""

    for row in source.iter_rows(min_row=4, min_col=2, max_col=7):
        value = row[0].value.strip()
        if value in source_description:
            source_description[value] = row[1].value.replace(" ", "").lower().replace(".", "")
            true_description[value] = row[1].value
            source_resp[value] = row[-1].value

    for row in target.iter_rows(min_row=4, min_col=2, max_col=4):
        id = row[0].value
        if id:
            id = id.strip()
            comp_desc = row[-1].value.replace(" ", "").lower().replace(".", "")
            if id in source_description:
                desc = source_description[id]
                if desc == comp_desc:
                    print(True)
                    target['I{}'.format(row[0].row)] = true_description[id]
                    target['J{}'.format(row[0].row)] = source_resp[id]

            else:
                raise Exception()

save_wbs()



